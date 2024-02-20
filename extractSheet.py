import os
import openpyxl
import sys
import subprocess
import platform
from tkinter import *
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import threading
import time


output_folder_path = r"uespoirSheet_sheets"
input_workbook_path = ""


def open_file_explorer():
    global input_path
    # Open a file dialog to select a file
    file_path = filedialog.askopenfilename()

    # Check if a file was selected
    if file_path:
        print("Selected file:", file_path)
        input_workbook_path = r"{}".format(file_path)
        thread = threading.Thread(target=save_sheets_to_folder(input_workbook_path, output_folder_path))
        thread.start()

    else:
        print("No file selected")



def save_sheets_to_folder(workbook_path, output_folder):

    messagebox.showinfo("Task Started", "Task Started, Press Ok and Wait...")

    # Load the Excel workbook
    workbook = openpyxl.load_workbook(workbook_path)
    message_label.config(text="Action Started, Wait...")

    # Create the output folder if it doesn't exist
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Iterate through all sheets in the workbook
    for sheet_name in workbook.sheetnames:
        # Create a new workbook for the current sheet
        new_workbook = openpyxl.Workbook()
        new_worksheet = new_workbook.active
        new_worksheet.title = sheet_name

        # Copy the contents and formatting of the current sheet to the new workbook
        source_sheet = workbook[sheet_name]
        for row in source_sheet:
            for cell in row:
                new_cell = new_worksheet.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
                new_cell.font = cell.font.copy()
                new_cell.border = cell.border.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.number_format = cell.number_format
                new_cell.protection = cell.protection.copy()
                new_cell.alignment = cell.alignment.copy()

        # Save the new workbook as a separate file
        new_workbook.save(os.path.join(output_folder, f"{sheet_name}.xlsx"))
        message_label.config(text="Action Done")


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)



# Creation of the window
window = tk.Tk()
window.geometry("700x500")
window.title("Tous les cours")
window.iconbitmap(resource_path('logo.ico'))

# show a label
Txt_label = tk.Label(window, text="Cliquez le bouton Upload pou choisir le fichier. ", font=("Helvetica", 16))
Txt_label.pack(pady=50)


Upload_button = tk.Button(window, text='Upload', width=15, height=1, bg='#156838', font="Lora", fg='white', command=open_file_explorer)
Upload_button.pack(pady=50)

# show a label
message_label = tk.Label(window, text="", font=("Helvetica", 12))
message_label.pack(pady=10)

window.mainloop()